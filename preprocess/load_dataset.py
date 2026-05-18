import os
import sys
currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)
sys.path.append(parentdir)

import numpy as np
import pandas as pd
import zipfile
import urllib.request
from openpyxl import load_workbook
from set_log import set_logging
####
# TableS4A - Whole set of log(IC50s) across all the screened compounds and cell lines
####
# GEX is short for gene expression
# CNV is short for copy number variation
# MET is short for DNA methylation
# WES is short for whole genome sequence


def main():
    NAME = "GDSC_dataloader"
    logger = set_logging(NAME)

    data_dir = os.getcwd()+"/data/GDSC_ALL"
    print(data_dir)
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)

    # Download raw files only if not already present locally.
    gex_file_check = os.path.join(data_dir, 'Cell_line_RMA_proc_basalExp.txt')
    if os.path.exists(gex_file_check):
        logger.info("Raw data files already present in {}, skipping download.".format(data_dir))
    else:
        urls_file = os.path.join(os.getcwd(), "preprocess", "load_GDSC.txt")
        urls = []
        with open(urls_file) as f:
            for line in f:
                if line.startswith("http://") or line.startswith("https://"):
                    urls.append(line.strip())

        logger.info("Downloading raw data from URLs...")
        for url in urls:
            local_file = os.path.join(data_dir, os.path.basename(url))
            if os.path.exists(local_file):
                continue
            urllib.request.urlretrieve(url, local_file)
            if local_file.endswith('.zip'):
                with zipfile.ZipFile(local_file, mode='r') as r:
                    r.extractall(data_dir)

    print("processing the GDSC data set")

    # --- READ GEX ---
    print("gene expression data")
    GEX_file = os.path.join(data_dir, 'Cell_line_RMA_proc_basalExp.txt')
    GEX = pd.read_csv(GEX_file, sep='\t')
    GEX_gene_symbols = np.array(GEX['GENE_SYMBOLS'], dtype='str')
    GEX = GEX.drop(['GENE_SYMBOLS', 'GENE_title'], axis=1)
    GEX_cell_ids = np.array(GEX.columns, dtype='str')
    for i, cell_id in enumerate(GEX_cell_ids):
        GEX_cell_ids[i] = cell_id[5:]
    GEX = np.array(GEX.values, dtype=float).T  # (1018,17737)
    logger.info("the GEX data has dims {}, {}".format(GEX.shape[0], GEX.shape[1]))

    # --- READ WES (optional) ---
    WES_cell_ids = None
    WES_CG = None
    WES_file = os.path.join(data_dir, 'CellLines_CG_BEMs/PANCAN_SEQ_BEM.txt')
    if os.path.exists(WES_file):
        print('read whole exome sequencing data')
        WES = pd.read_csv(WES_file, sep='\t')
        WES_CG = np.array(WES['CG'], dtype='str')
        WES = WES.drop(['CG'], axis=1)
        WES_cell_ids = np.array(WES.columns, dtype='str')
        WES = np.array(WES.values, dtype=int).T  # (961,300)
        logger.info("the WES data has dims {}, {}".format(WES.shape[0], WES.shape[1]))
    else:
        logger.warning("WES file not found, skipping: {}".format(WES_file))
        WES = None

    # --- READ CNV (optional) ---
    CNV_cell_ids = None
    CNV_cna = None
    CNV_file = os.path.join(data_dir, 'CellLine_CNV_BEMs/PANCAN_CNA_BEM.rdata.txt')
    if os.path.exists(CNV_file):
        print("Read Copy number dataset")
        CNV = pd.read_csv(CNV_file, sep='\t')
        CNV_cell_ids = np.array(CNV['Unnamed: 0'], dtype='str')
        CNV = CNV.drop(['Unnamed: 0'], axis=1)
        CNV_cna = np.array(CNV.columns, dtype='str')
        CNV = np.array(CNV.values, dtype=int)  # (996,425)
        logger.info("the CNV data has dims {}, {}".format(CNV.shape[0], CNV.shape[1]))
    else:
        logger.warning("CNV file not found, skipping: {}".format(CNV_file))
        CNV = None

    # --- READ MET (optional) ---
    MET_cell_ids = None
    MET_met = None
    MET_file = os.path.join(data_dir, 'METH_CELLLINES_BEMs/PANCAN.txt')
    if os.path.exists(MET_file):
        print("Read Methylation dataset")
        MET = pd.read_csv(MET_file, sep='\t')
        MET_met = np.array(MET['Unnamed: 0'], dtype='str')
        MET = MET.drop(['Unnamed: 0'], axis=1)
        MET_cell_ids = np.array(MET.columns, dtype='str')
        MET = np.array(MET.values, dtype=int).T  # (790,378)
        logger.info("the MET data has dims {}, {}".format(MET.shape[0], MET.shape[1]))
    else:
        logger.warning("MET file not found, skipping: {}".format(MET_file))
        MET = None

    # --- READ IC50 ---
    print("Read LOG_IC50 dataset")
    print("all the IC50 from the xlsx file are log IC50")
    IC50_file = os.path.join(data_dir, 'TableS4A.xlsx')
    wb = load_workbook(filename=IC50_file)
    sheet = wb['TableS4A-IC50s']
    IC50_cell_ids, IC50_cell_names = [], []
    for i in range(7, 997):
        IC50_cell_ids.append('%s' % sheet['A%s' % i].value)
        IC50_cell_names.append(('%s' % sheet['B%s' % i].value).strip())
    IC50_cell_ids = np.array(IC50_cell_ids, dtype='str')
    IC50_cell_names = np.array(IC50_cell_names, dtype='str')
    logger.info("Log IC50 originally have cells {}".format(IC50_cell_ids.shape[0]))

    IC50_drug_ids, IC50_drug_names = [], []
    for i, (cell_row5, cell_row6) in enumerate(zip(sheet[5], sheet[6])):
        if i > 1:
            IC50_drug_ids.append('%s' % cell_row5.value)
            IC50_drug_names.append(('%s' % cell_row6.value).strip())
    IC50_drug_ids = np.array(IC50_drug_ids, dtype='str')
    IC50_drug_names = np.array(IC50_drug_names, dtype='str')
    logger.info("Log IC50 originally have Drugs {}".format(IC50_drug_ids.shape[0]))

    IC50 = np.ones([IC50_cell_ids.shape[0], IC50_drug_ids.shape[0]]) * np.nan  # n*m
    for i in range(7, 997):
        for j, cell in enumerate(sheet[i]):
            if j > 1:
                if cell.value != 'NA':
                    IC50[i - 7, j - 2] = cell.value

    # Read LOG_IC50 Threshold
    threshold_file = os.path.join(data_dir, 'TableS5C.xlsx')
    wb = load_workbook(filename=threshold_file)
    sheet = wb['Table-S5C binaryIC50s']
    threshold = []
    for i, cell in enumerate(sheet[7]):
        if i > 1:
            threshold.append(cell.value)
    threshold = np.array(threshold)
    drug_ids_file = os.path.join(data_dir, 'TableS1F.xlsx')
    wb = load_workbook(filename=drug_ids_file)
    sheet = wb['TableS1F_ScreenedCompounds']
    threshold_drug_ids = []
    for i in range(4, 269):
        threshold_drug_ids.append('%s' % sheet['B%s' % i].value)
    threshold_drug_ids = np.array(threshold_drug_ids)
    logger.info("threshold drugs have {}".format(threshold_drug_ids.shape[0]))

    # Normalize IC50 by the threshold
    merged = intersect_index(IC50_drug_ids, threshold_drug_ids)
    IC50_keep_index = np.array(merged['index1'].values, dtype=int)
    IC50_drug_ids = IC50_drug_ids[IC50_keep_index]
    IC50 = IC50[:, IC50_keep_index]
    threshold_keep_index = np.array(merged['index2'].values, dtype=int)
    threshold = threshold[threshold_keep_index]

    # this is just -logIC50+log(threshold), max 8.125, min -11.47
    IC50_norm = - (IC50 - threshold)
    IC50_norm_min = np.min(IC50_norm[~np.isnan(IC50_norm)])  # -11.47
    IC50_norm = IC50_norm - IC50_norm_min  # min 0.0, max 19.6
    logger.info(
        "after normalize, the max LogIC50 is {:.5f} and min LogIC50 is {:.4f}".format(
            np.nanmax(IC50_norm), np.nanmin(IC50_norm)))

    # --- SAVE GEX ---
    merged = intersect_index(GEX_cell_ids, IC50_cell_ids)
    GEX_keep_index = np.array(merged['index1'].values, dtype=int)
    IC50_keep_index = np.array(merged['index2'].values, dtype=int)
    GEX = GEX[GEX_keep_index]
    GEX_cell_ids = GEX_cell_ids[GEX_keep_index]
    GEX_cell_names = IC50_cell_names[IC50_keep_index]
    IC50_gex = IC50_norm[IC50_keep_index]
    np.savez('%s/GDSC_GEX.npz' % data_dir, X=GEX, Y=IC50_gex, cell_ids=GEX_cell_ids,
             cell_names=GEX_cell_names, drug_ids=IC50_drug_ids, drug_names=IC50_drug_names,
             GEX_gene_symbols=GEX_gene_symbols)
    logger.info('Gene expression (GEX) dataset: {} cell lines, {} features, {} drugs'.format(
        GEX.shape[0], GEX.shape[1], IC50_gex.shape[1]))

    # --- SAVE WES (if available) ---
    if WES is not None:
        merged = intersect_index(WES_cell_ids, IC50_cell_ids)
        WES_keep_index = np.array(merged['index1'].values, dtype=int)
        IC50_keep_index = np.array(merged['index2'].values, dtype=int)
        WES = WES[WES_keep_index]
        WES_cell_ids = WES_cell_ids[WES_keep_index]
        WES_cell_names = IC50_cell_names[IC50_keep_index]
        np.savez('%s/GDSC_WES.npz' % data_dir, X=WES, Y=IC50_norm[IC50_keep_index],
                 cell_ids=WES_cell_ids, cell_names=WES_cell_names,
                 drug_ids=IC50_drug_ids, drug_names=IC50_drug_names, WES_CG=WES_CG)
        logger.info('Whole-exome sequencing (WES) dataset: {} cell lines, {} features, {} drugs'.format(
            WES.shape[0], WES.shape[1], IC50_norm.shape[1]))

    # --- SAVE CNV (if available) ---
    if CNV is not None:
        merged = intersect_index(CNV_cell_ids, IC50_cell_ids)
        CNV_keep_index = np.array(merged['index1'].values, dtype=int)
        IC50_keep_index = np.array(merged['index2'].values, dtype=int)
        CNV = CNV[CNV_keep_index]
        CNV_cell_ids = CNV_cell_ids[CNV_keep_index]
        CNV_cell_names = IC50_cell_names[IC50_keep_index]
        np.savez('%s/GDSC_CNV.npz' % data_dir, X=CNV, Y=IC50_norm[IC50_keep_index],
                 cell_ids=CNV_cell_ids, cell_names=CNV_cell_names,
                 drug_ids=IC50_drug_ids, drug_names=IC50_drug_names, CNV_cna=CNV_cna)
        logger.info('Copy number variation (CNV) dataset: {} cell lines, {} features, {} drugs'.format(
            CNV.shape[0], CNV.shape[1], IC50_norm.shape[1]))

    # --- SAVE MET (if available) ---
    if MET is not None:
        merged = intersect_index(MET_cell_ids, IC50_cell_ids)
        MET_keep_index = np.array(merged['index1'].values, dtype=int)
        IC50_keep_index = np.array(merged['index2'].values, dtype=int)
        MET = MET[MET_keep_index]
        MET_cell_ids = MET_cell_ids[MET_keep_index]
        MET_cell_names = IC50_cell_names[IC50_keep_index]
        np.savez('%s/GDSC_MET.npz' % data_dir, X=MET, Y=IC50_norm[IC50_keep_index],
                 cell_ids=MET_cell_ids, cell_names=MET_cell_names,
                 drug_ids=IC50_drug_ids, drug_names=IC50_drug_names, MET_met=MET_met)
        logger.info('Methylation (MET) dataset: {} cell lines, {} features, {} drugs'.format(
            MET.shape[0], MET.shape[1], IC50_norm.shape[1]))

    print('Done')


def intersect_index(list1, list2):
    '''
    Given two lists find the index of intersect in list1

    Parameters
    ----------
    list1: 1d numpy array
    list2: 1d numpy array
    '''
    intersect = np.intersect1d(list1, list2)
    intersect = pd.DataFrame(intersect, columns=['id'])
    list1 = np.vstack([np.arange(list1.shape[0]), list1]).T
    list1 = pd.DataFrame(list1, columns=['index1', 'id'])
    list2 = np.vstack([np.arange(list2.shape[0]), list2]).T
    list2 = pd.DataFrame(list2, columns=['index2', 'id'])
    merged = pd.merge(list1, intersect, on='id', how='right')
    merged = pd.merge(merged, list2, on='id', how='left')

    return merged


if __name__ == '__main__':
    main()
